# -*- coding: utf-8 -*-
"""AI总结模块 — DeepSeek-V4-Flash直接提取PDF文字(无需OCR)
稳定性优先: 每个模板最多重试5次, 确保全部执行
用法: python ai_summarize.py --date 2026-06-23 --skip-trend
环境变量: SILICONFLOW_API_KEY
"""

import os, re, json, time, argparse, zipfile
import urllib.request, urllib.error
from datetime import datetime

# ═══ Config ═══
API_URL = "https://api.siliconflow.cn/v1/chat/completions"
MODEL_TEXT = "deepseek-ai/DeepSeek-V4-Flash"
TIMEOUT_API = 90
MAX_ATT_TEXT = 5000
MAX_INPUT_CHARS = 8000
MAX_RETRIES = 5
RETRY_DELAY = 5

from keywords import PROMPT_AWARD_INSIGHT, PROMPT_MARKET_INTEL, PROMPT_TREND_RADAR

NO_BOILERPLATE = "\n\n重要约束:\n1. 全部用中文回答\n2. 严禁出现我是、作为一名、分析师等自我介绍套话\n3. 严禁出现与中标信息无关的内容\n4. 直接输出分析结果,不要寒暄\n5. 输出格式要求: 直接使用HTML标签(如<h2>标题</h2>, <p>段落</p>, <ul><li>列表项</li></ul>, <table><tr><td>表格</td></tr></table>, <strong>加粗</strong>), 不要使用Markdown语法(不要用#、-、**等标记)\n6. 严禁重复输出相同内容,每个章节只输出一次,不要复制粘贴同样的句子"

def log(*a):
    print(*a, flush=True)

def call_api(messages, max_tokens=2000, retries=MAX_RETRIES):
    api_key = os.environ.get("SILICONFLOW_API_KEY", "")
    if not api_key:
        return "[ERROR] SILICONFLOW_API_KEY 未设置"

    for attempt in range(retries):
        try:
            payload = json.dumps({
                "model": MODEL_TEXT,
                "messages": messages,
                "max_tokens": max_tokens,
                "temperature": 0.3
            }).encode("utf-8")

            req = urllib.request.Request(API_URL, data=payload, method="POST")
            req.add_header("Authorization", f"Bearer {api_key}")
            req.add_header("Content-Type", "application/json")

            with urllib.request.urlopen(req, timeout=TIMEOUT_API) as resp:
                result = json.loads(resp.read().decode("utf-8"))
                content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
                if content and len(content) > 10:
                    return content
                log(f"    API返回空, 重试({attempt+1}/{retries})")
        except Exception as e:
            err = str(e)[:80]
            if attempt < retries - 1:
                wait = RETRY_DELAY * (attempt + 1)
                log(f"    API错误({attempt+1}/{retries}): {err}, 等待{wait}s...")
                time.sleep(wait)
            else:
                return f"[API错误] {err}"
    return "[API错误] 重试次数耗尽"

# ═══ 附件文本提取 — 纯代码,无OCR ═══
def extract_pdf_text(pdf_path):
    try:
        import pdfplumber
        pages_text = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:10]:
                t = page.extract_text() or ""
                pages_text.append(t)
        full_text = "\n".join(pages_text).strip()
        return full_text[:MAX_ATT_TEXT] if full_text else ""
    except Exception as e:
        log(f"    PDF提取失败: {str(e)[:50]}")
        return ""

def extract_docx_text(docx_path):
    try:
        from docx import Document
        doc = Document(docx_path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())[:MAX_ATT_TEXT]
    except:
        return ""

def extract_xlsx_text(xlsx_path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(max_row=100, values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append(" | ".join(cells))
        return "\n".join(lines)[:MAX_ATT_TEXT]
    except:
        return ""

def extract_attachment_text(file_path):
    if not os.path.exists(file_path):
        return ""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return extract_pdf_text(file_path)
    elif ext in (".doc", ".docx"):
        return extract_docx_text(file_path)
    elif ext in (".xls", ".xlsx"):
        return extract_xlsx_text(file_path)
    return ""

def get_attachment_files(out_dir, dir_name):
    att_dir = os.path.join(out_dir, dir_name)
    if not os.path.isdir(att_dir):
        return []
    files = []
    for f in os.listdir(att_dir):
        fp = os.path.join(att_dir, f)
        if os.path.isfile(fp) and not f.startswith(".") and f.lower().endswith((".pdf", ".doc", ".docx", ".xls", ".xlsx")):
            files.append(fp)
    return files[:5]

def clean_title_for_dir(title):
    t = re.sub(r'[\\/:*?"<>|]', "_", title or "")
    t = re.sub(r"\s+", " ", t).strip()
    return t[:80] if t else "untitled"

def wrap_html(title, template_name, content):
    css = "body{font-family:Microsoft YaHei,sans-serif;max-width:900px;margin:40px auto;padding:20px;line-height:1.8;color:#333}h1{color:#1a5276;border-bottom:3px solid #2980b9;padding-bottom:10px}h2{color:#2980b9;margin-top:30px}h3{color:#34495e}table{border-collapse:collapse;width:100%;margin:15px 0}th,td{border:1px solid #ddd;padding:8px 12px;text-align:left}th{background-color:#ebf5fb}strong{color:#c0392b}hr{border:none;border-top:2px solid #eee;margin:30px 0}"
    return f"""<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><title>{title} - {template_name}</title><style>{css}</style></head><body><h1>{title}</h1><h2>{template_name}</h2><hr>{content}</body></html>"""

def process_item(item, out_dir):
    title = item.get("标题", "")
    url = item.get("原文链接", "")

    dir_name = clean_title_for_dir(title)
    att_files = get_attachment_files(out_dir, dir_name)
    att_text = ""
    if att_files:
        log(f"    提取{len(att_files)}个附件文字...")
        for fp in att_files:
            fname = os.path.basename(fp)
            text = extract_attachment_text(fp)
            if text and len(text) > 10:
                att_text += f"\n\n--- 附件: {fname} ---\n{text}"

    announcement_text = f"""
标题: {title}
采购人: {item.get("采购人", "")}
代理机构: {item.get("代理机构", "")}
供应商: {item.get("供应商", "")}
供应商地址: {item.get("供应商地址", "")}
中标金额(万): {item.get("中标金额(万)", "")}
公告类型: {item.get("公告类型", "")}
地区: {item.get("地区", "")}
发布时间: {item.get("发布时间", "")}
原文链接: {url}
"""
    full_input = announcement_text + att_text
    if len(full_input) > MAX_INPUT_CHARS:
        full_input = full_input[:MAX_INPUT_CHARS] + "\n...(内容截断)"

    awards_dir = os.path.join(out_dir, "AI总结")
    os.makedirs(awards_dir, exist_ok=True)
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", title)[:50]

    # 模板1: Award Insight (文件名: 标准分析_标题.html)
    log(f"    [模板1] 标准分析...")
    r1 = call_api([{"role": "user", "content": PROMPT_AWARD_INSIGHT + full_input + NO_BOILERPLATE}], max_tokens=2000)
    with open(os.path.join(awards_dir, f"标准分析_{safe_name}.html"), "w", encoding="utf-8") as f:
        f.write(wrap_html(title, "CHD-Award Insight 标准分析", r1))
    log(f"    [模板1] 完成: {len(r1)}字")

    # 模板2: Market Intelligence (文件名: 市场情报_标题.html)
    log(f"    [模板2] 市场情报...")
    r2 = call_api([{"role": "user", "content": PROMPT_MARKET_INTEL + full_input + NO_BOILERPLATE}], max_tokens=2000)
    with open(os.path.join(awards_dir, f"市场情报_{safe_name}.html"), "w", encoding="utf-8") as f:
        f.write(wrap_html(title, "CHD-Market Intelligence 市场情报", r2))
    log(f"    [模板2] 完成: {len(r2)}字")

    # 提取Executive Summary
    exec_summary = ""
    for idx, line in enumerate(r1.split("\n")):
        if "Executive Summary" in line or "执行摘要" in line or "6." in line[:5]:
            for nl in r1.split("\n")[idx+1:]:
                nl = nl.strip()
                if nl and not nl.startswith("#") and not nl.startswith("<") and not nl.startswith("---"):
                    exec_summary = nl
                    break
            if exec_summary:
                break
    if not exec_summary:
        exec_summary = r1[-300:].strip()

    return {
        "标题": title,
        "AwardInsight": r1,
        "MarketIntel": r2,
        "ExecSummary": exec_summary[:300]
    }

def trend_radar(summaries, out_dir):
    log("\n[*] 模板3: 行业趋势分析...")
    input_parts = []
    for i, s in enumerate(summaries):
        ai_text = s.get("AwardInsight", "")
        if ai_text.startswith("["):
            ai_text = s.get("标题", "")
        input_parts.append(f"--- 公告{i+1} ---\n标题: {s['标题']}\n分析:\n{ai_text[:500]}")
    input_text = "\n\n".join(input_parts)
    if len(input_text) > MAX_INPUT_CHARS:
        input_text = input_text[:MAX_INPUT_CHARS] + "\n...(内容截断)"

    result = call_api([{"role": "user", "content": PROMPT_TREND_RADAR + input_text + NO_BOILERPLATE}], max_tokens=2500)
    with open(os.path.join(out_dir, "行业趋势分析.html"), "w", encoding="utf-8") as f:
        f.write(wrap_html("文物数字化行业趋势分析", "CHD-Trend Radar 行业趋势分析", result))
    log(f"  模板3完成: {len(result)}字")
    return result

def run(date_str=None, skip_trend=False):
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")

    base = os.environ.get("DATA_BASE", "")
    out_dir = os.path.join(base, "output", date_str) if base else os.path.join("output", date_str)

    json_path = os.path.join(out_dir, "文物数字化.json")

    if not os.path.exists(json_path):
        jsons = [f for f in os.listdir(out_dir) if f.endswith(".json") and not any(x in f.lower() for x in ["attach", "summarie"])] if os.path.isdir(out_dir) else []
        if jsons:
            json_path = os.path.join(out_dir, jsons[0])
        else:
            log(f"[!] 未找到JSON: {json_path}")
            return

    with open(json_path, "r", encoding="utf-8") as f:
        items = json.load(f)

    log(f"[*] 读取 {len(items)} 条公告, 开始AI总结(纯代码提取, 无OCR)")
    log(f"    模型: {MODEL_TEXT}")
    log(f"    重试: 最多{MAX_RETRIES}次/模板")
    log(f"    输出: HTML格式\n")

    t0 = time.time()
    summaries = []

    for i, item in enumerate(items, 1):
        title = item.get("标题", "")
        log(f"\n  [{i}/{len(items)}] {title[:40]}...")
        try:
            s = process_item(item, out_dir)
            summaries.append(s)
            status = "OK" if not s["AwardInsight"].startswith("[") else "FAIL"
            log(f"  [{i}/{len(items)}] {status}")
        except Exception as e:
            log(f"  [{i}/{len(items)}] FAIL: {str(e)[:60]}")
            summaries.append({"标题": title, "AwardInsight": f"[错误]{str(e)[:60]}", "MarketIntel": "", "ExecSummary": ""})

    if summaries and not skip_trend:
        trend_radar(summaries, out_dir)

    summaries_path = os.path.join(out_dir, "summaries.json")
    with open(summaries_path, "w", encoding="utf-8") as f:
        json.dump(summaries, f, ensure_ascii=False, indent=2)

    ok_count = sum(1 for s in summaries if not s.get("AwardInsight", "").startswith("["))
    log(f"\n[*] AI总结完成: {ok_count}/{len(summaries)}成功, 耗时{time.time()-t0:.0f}s")

if __name__ == "__main__":
    p = argparse.ArgumentParser(description="AI总结模块")
    p.add_argument("--date", default=None, help="日期 YYYY-MM-DD")
    p.add_argument("--skip-trend", action="store_true", help="跳过模板3行业趋势分析")
    a = p.parse_args()
    run(a.date, a.skip_trend)
